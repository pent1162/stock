#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Research Report Parser (研報解析腳本)
======================================
Extracts structured investment information from PDF research reports.
Supports Chinese and English reports, scanned PDFs (OCR), and batch processing.

Outputs:
  - JSON:  /home/sprite/quant/output/parsed_report_{filename}.json
  - TXT:   /home/sprite/quant/output/parsed_report_{filename}.txt
  - XLSX:  /home/sprite/quant/output/report_tracker.xlsx  (cumulative)

Usage:
  python research_parser.py                          # batch: all PDFs in reports/
  python research_parser.py path/to/report.pdf       # single file
  python research_parser.py --demo                   # run demo with fake data
"""

import os, re, sys, json, logging
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR    = Path("/home/sprite/quant")
REPORTS_DIR = BASE_DIR / "reports"
OUTPUT_DIR  = BASE_DIR / "output"
TRACKER     = OUTPUT_DIR / "report_tracker.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# 1. PDF TEXT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

def extract_text_pdfplumber(pdf_path: str) -> str:
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            pages = [p.extract_text() or "" for p in pdf.pages]
        return "\n".join(pages)
    except Exception as e:
        log.warning(f"pdfplumber failed: {e}")
        return ""

def extract_text_pymupdf(pdf_path: str) -> str:
    try:
        import fitz
        doc = fitz.open(pdf_path)
        pages = [doc[i].get_text() for i in range(len(doc))]
        doc.close()
        return "\n".join(pages)
    except Exception as e:
        log.warning(f"PyMuPDF failed: {e}")
        return ""

def extract_text_ocr(pdf_path: str) -> str:
    try:
        import fitz, pytesseract
        from PIL import Image
        import io
        doc = fitz.open(pdf_path)
        texts = []
        for i in range(len(doc)):
            pix = doc[i].get_pixmap(dpi=200)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            txt = pytesseract.image_to_string(img, lang="chi_tra+chi_sim+eng")
            texts.append(txt)
        doc.close()
        log.info(f"OCR extracted {sum(len(t) for t in texts)} chars")
        return "\n".join(texts)
    except Exception as e:
        log.warning(f"OCR failed: {e}")
        return ""

def extract_text(pdf_path: str) -> str:
    for fn in (extract_text_pdfplumber, extract_text_pymupdf):
        text = fn(pdf_path)
        if text and len(text.strip()) > 100:
            return text
    log.info("Digital extraction yielded little text — trying OCR...")
    return extract_text_ocr(pdf_path)


# ══════════════════════════════════════════════════════════════════════════════
# 2. FIELD EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════

# ── 2a. Target Price ──────────────────────────────────────────────────────────
_TP_PATTERNS = [
    r'(?:目標價(?:位)?|目标价(?:位)?)\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|US\\$?|RMB|人民幣|台幣|港幣|美元|\\$|¥)?\\s*([\\d,，.]+)',
    r'[Tt]arget\\s+[Pp]rice\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|US\\$?|RMB|\\$|¥)?\\s*([\\d,，.]+)',
    r'\\b[Pp][Tt]\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|US\\$?|\\$)?\\s*([\\d,，.]+)',
    r'上調目標價至\\s*(?:NT\\$?|HK\\$?|\\$)?\\s*([\\d,，.]+)',
    r'下調目標價至\\s*(?:NT\\$?|HK\\$?|\\$)?\\s*([\\d,，.]+)',
]

def extract_target_price(text: str) -> Optional[str]:
    for pat in _TP_PATTERNS:
        m = re.search(pat, text)
        if m:
            raw = m.group(1).replace('，', '').strip()
            try:
                float(raw.replace(',', ''))
                return raw
            except ValueError:
                continue
    return None


# ── 2b. Rating ────────────────────────────────────────────────────────────────
_RATING_MAP = {
    '強烈買入': 'STRONG BUY', '强烈买入': 'STRONG BUY',
    '買入':     'BUY',        '买入':     'BUY',
    '增持':     'OVERWEIGHT',
    '優於大市': 'OUTPERFORM', '优于大市': 'OUTPERFORM',
    '跑贏大市': 'OUTPERFORM', '跑赢大市': 'OUTPERFORM',
    '中性':     'NEUTRAL',
    '持有':     'HOLD',
    '減持':     'UNDERWEIGHT','减持':     'UNDERWEIGHT',
    '跑輸大市': 'UNDERPERFORM','跑输大市': 'UNDERPERFORM',
    '賣出':     'SELL',       '卖出':     'SELL',
    'STRONG BUY': 'STRONG BUY', 'BUY': 'BUY',
    'OUTPERFORM': 'OUTPERFORM', 'OVERWEIGHT': 'OVERWEIGHT',
    'NEUTRAL': 'NEUTRAL', 'HOLD': 'HOLD',
    'UNDERPERFORM': 'UNDERPERFORM', 'UNDERWEIGHT': 'UNDERWEIGHT',
    'SELL': 'SELL',
}

def extract_rating(text: str) -> Optional[str]:
    label_pat = r'(?:投資評級|評級|评级|Rating)\\s*[：:=]?\\s*(' + '|'.join(re.escape(k) for k in _RATING_MAP) + r')'
    m = re.search(label_pat, text, re.IGNORECASE)
    if m:
        raw = m.group(1).strip()
        return _RATING_MAP.get(raw, _RATING_MAP.get(raw.upper(), raw.upper()))
    bare_pat = r'\\b(' + '|'.join(re.escape(k) for k in _RATING_MAP if len(k) > 1) + r')\\b'
    m = re.search(bare_pat, text, re.IGNORECASE)
    if m:
        raw = m.group(1).strip()
        return _RATING_MAP.get(raw, _RATING_MAP.get(raw.upper(), raw.upper()))
    return None


# ── 2c. Key Financial Metrics ─────────────────────────────────────────────────
_METRIC_PATTERNS = {
    'EPS': [
        r'EPS\\s*(?:預測|预测|forecast|estimate)?\\s*(?:[12]\\d{3}[EFe])?\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|\\$)?\\s*([\\d.]+)',
        r'每股(?:盈利|收益)\\s*[：:=]?\\s*([\\d.]+)',
    ],
    'PE': [
        r'P/?E\\s*(?:[12]\\d{3}[EFe])?\\s*[：:=]?\\s*([\\d.]+)\\s*[xX倍]?',
        r'市盈率\\s*[：:=]?\\s*([\\d.]+)',
    ],
    'ROE': [
        r'ROE\\s*[：:=]?\\s*([\\d.]+)\\s*%?',
        r'股本回報率\\s*[：:=]?\\s*([\\d.]+)',
    ],
    'Revenue': [
        r'(?:Revenue|Revenues|營收|营收|收入)\\s*(?:[12]\\d{3}[EFe])?\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|US\\$?|RMB|\\$|¥)?\\s*([\\d,.]+\\s*(?:億|百萬|兆|trillion|billion|million|[BMbm])+)',
        r'預計收入\\s*[：:=]?\\s*(?:NT\\$?|HK\\$?|\\$)?\\s*([\\d,.]+\\s*(?:億|百萬|兆|billion|million)?)',
    ],
    'Gross_Margin': [
        r'(?:毛利率|Gross\\s+Margin)\\s*[：:=]?\\s*([\\d.]+)\\s*%?',
    ],
    'Net_Margin': [
        r'(?:淨利率|净利率|Net\\s+(?:Profit\\s+)?Margin)\\s*[：:=]?\\s*([\\d.]+)\\s*%?',
    ],
}

def extract_metrics(text: str) -> dict:
    metrics = {}
    for key, patterns in _METRIC_PATTERNS.items():
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                metrics[key] = m.group(1).strip()
                break
    return metrics


# ── 2d. Investment Thesis ─────────────────────────────────────────────────────
_THESIS_ANCHORS = [
    '核心觀點', '核心观点', '投資邏輯', '投资逻辑', '核心投資邏輯',
    '投資亮點', '投资亮点', '核心邏輯', '核心逻辑',
    'Investment Thesis', 'Key Investment', 'Key Points', 'Highlights',
    '我們認為', '我们认为', '主要觀點', '主要观点',
]
_BULLET_RE = re.compile(
    r'^\s*(?:[①②③④⑤⑥⑦⑧⑨⑩]|[1-9][.、。)]|[-•\*]|[一二三四五六七八九十][、。])',
    re.MULTILINE,
)

def extract_thesis(text: str) -> list:
    for anchor in _THESIS_ANCHORS:
        idx = text.find(anchor)
        if idx == -1:
            continue
        segment = text[idx: idx + 1500]
        bullets = _BULLET_RE.split(segment)
        points = [b.strip().split('\n')[0].strip() for b in bullets if len(b.strip()) > 15]
        if points:
            return points[:3]
    sentences = re.split(r'[。！？.!?]', text)
    return [s.strip() for s in sentences if len(s.strip()) > 20][:3]


# ── 2e. Risk Factors ──────────────────────────────────────────────────────────
_RISK_ANCHORS = [
    '風險提示', '风险提示', '風險因素', '风险因素',
    '主要風險', '主要风险', '潛在風險', '潜在风险',
    'Risk Factors', 'Key Risks', 'Risks', 'Risk Warning',
]

def extract_risks(text: str) -> list:
    for anchor in _RISK_ANCHORS:
        idx = text.find(anchor)
        if idx == -1:
            continue
        segment = text[idx: idx + 1000]
        bullets = _BULLET_RE.split(segment)
        risks = [b.strip().split('\n')[0].strip() for b in bullets if len(b.strip()) > 10]
        if risks:
            return risks[:5]
        sentences = re.split(r'[。！？.!?]', segment)
        return [s.strip() for s in sentences if len(s.strip()) > 10][:5]
    return []


# ── 2f. Date ──────────────────────────────────────────────────────────────────
def extract_date(text: str) -> Optional[str]:
    patterns = [
        r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})[日]?',
        r'(\d{4})[./](\d{2})[./](\d{2})',
    ]
    for pat in patterns:
        m = re.search(pat, text[:3000])
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


# ── 2g. Institution & Analyst ─────────────────────────────────────────────────
_INSTITUTION_PATTERNS = [
    r'(?:機構|机构|出具方|發布方|发布方|Issuer|Firm|Bank|Securities|Research)\s*[：:=]?\s*([^\n,，。]{2,30})',
    r'^([^\n]{2,20}(?:證券|证券|資本|资本|Securities|Capital|Research|Invest|Bank|Asset))',
]
_ANALYST_PATTERNS = [
    r'(?:分析師|分析师|研究員|研究员|Analyst|Author|Written\s+by)\s*[：:=]?\s*([^\n,，。]{2,20})',
    r'撰寫人\s*[：:=]?\s*([^\n,，。]{2,15})',
]

def _first_match(patterns: list, text: str, n: int = 2000) -> Optional[str]:
    snippet = text[:n]
    for pat in patterns:
        m = re.search(pat, snippet, re.MULTILINE)
        if m:
            return m.group(1).strip()
    return None

def extract_institution(text: str) -> Optional[str]:
    return _first_match(_INSTITUTION_PATTERNS, text)

def extract_analyst(text: str) -> Optional[str]:
    return _first_match(_ANALYST_PATTERNS, text)


# ── 2h. Ticker ────────────────────────────────────────────────────────────────
def extract_ticker(text: str) -> Optional[str]:
    snippet = text[:2000]
    m = re.search(r'([\u4e00-\u9fff]{2,8})\s*[（(]([\d]{4,6}(?:\.[A-Za-z]{2})?)[）)]', snippet)
    if m:
        return f"{m.group(1)} ({m.group(2)})"
    m = re.search(r'([A-Z][A-Za-z]{1,8})\s*[（(]([\d]{4,6}(?:\.[A-Za-z]{2})?)[）)]', snippet)
    if m:
        return f"{m.group(1)} ({m.group(2)})"
    m = re.search(r'\b(\d{4,6}\.[A-Z]{2})\b', snippet)
    if m:
        return m.group(1)
    m = re.search(r'\b([A-Z]{2,5})(?=\s*(?:Inc|Corp|Ltd|Co\.?))', snippet)
    if m:
        return m.group(1)
    return None


# ══════════════════════════════════════════════════════════════════════════════
# 3. MAIN PARSE FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def parse_report(pdf_path: str) -> dict:
    path = Path(pdf_path)
    log.info(f"Parsing: {path.name}")
    text = extract_text(pdf_path)
    if not text.strip():
        log.error("No text extracted — cannot parse.")
        return {}
    return {
        "source_file":    path.name,
        "parsed_at":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date":           extract_date(text),
        "institution":    extract_institution(text),
        "analyst":        extract_analyst(text),
        "ticker":         extract_ticker(text),
        "rating":         extract_rating(text),
        "target_price":   extract_target_price(text),
        "metrics":        extract_metrics(text),
        "thesis":         extract_thesis(text),
        "risks":          extract_risks(text),
        "raw_text_chars": len(text),
    }


# ══════════════════════════════════════════════════════════════════════════════
# 4. OUTPUT WRITERS
# ══════════════════════════════════════════════════════════════════════════════

def _rating_arrow(rating: Optional[str]) -> str:
    if not rating: return ""
    r = rating.upper()
    if r in ('BUY', 'STRONG BUY', 'OUTPERFORM', 'OVERWEIGHT'): return " ▲"
    if r in ('SELL', 'UNDERPERFORM', 'UNDERWEIGHT'): return " ▼"
    return " ─"

def write_json(data: dict, stem: str) -> Path:
    out = OUTPUT_DIR / f"parsed_report_{stem}.json"
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info(f"JSON  → {out}")
    return out

def write_txt(data: dict, stem: str) -> Path:
    d = data
    metrics = d.get('metrics', {})
    lines = [
        "📊 研報解析結果",
        "─────────────────",
        f"🏢 機構：{d.get('institution') or '(未識別)'}",
        f"👤 分析師：{d.get('analyst') or '(未識別)'}",
        f"📅 日期：{d.get('date') or '(未識別)'}",
        f"📈 標的：{d.get('ticker') or '(未識別)'}",
        f"⭐ 評級：{d.get('rating') or '(未識別)'}{_rating_arrow(d.get('rating'))}",
        f"🎯 目標價：{d.get('target_price') or '(未識別)'}",
    ]
    thesis = d.get('thesis', [])
    if thesis:
        lines += ["", "💡 核心投資邏輯："]
        for i, pt in enumerate(thesis, 1):
            lines.append(f"{i}. {pt[:120]}")
    if metrics:
        lines += ["", "📊 關鍵數據預測："]
        labels = {'EPS':'EPS','PE':'PE','ROE':'ROE','Revenue':'營收','Gross_Margin':'毛利率','Net_Margin':'淨利率'}
        for k, label in labels.items():
            if k in metrics:
                lines.append(f"• {label}：{metrics[k]}")
    risks = d.get('risks', [])
    if risks:
        lines += ["", "⚠️ 風險提示："]
        for r in risks:
            lines.append(f"• {r[:100]}")
    lines += ["─────────────────", f"🗂 來源：{d.get('source_file', '')}"]
    out = OUTPUT_DIR / f"parsed_report_{stem}.txt"
    out.write_text("\n".join(lines), encoding='utf-8')
    log.info(f"TXT   → {out}")
    return out

def write_tracker(data: dict) -> Path:
    HEADERS = [
        'Parsed At','Source File','Date','Institution','Analyst',
        'Ticker','Rating','Target Price',
        'EPS','PE','ROE','Revenue','Gross Margin','Net Margin',
        'Thesis 1','Thesis 2','Thesis 3',
        'Risk 1','Risk 2','Risk 3',
    ]
    HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
    ALT_FILL    = PatternFill('solid', fgColor='D6E4F0')

    metrics = data.get('metrics', {})
    thesis  = data.get('thesis', [])
    risks   = data.get('risks', [])

    new_row = [
        data.get('parsed_at',''), data.get('source_file',''), data.get('date',''),
        data.get('institution',''), data.get('analyst',''), data.get('ticker',''),
        data.get('rating',''), data.get('target_price',''),
        metrics.get('EPS',''), metrics.get('PE',''), metrics.get('ROE',''),
        metrics.get('Revenue',''), metrics.get('Gross_Margin',''), metrics.get('Net_Margin',''),
        thesis[0][:200] if len(thesis)>0 else '',
        thesis[1][:200] if len(thesis)>1 else '',
        thesis[2][:200] if len(thesis)>2 else '',
        risks[0][:200]  if len(risks)>0  else '',
        risks[1][:200]  if len(risks)>1  else '',
        risks[2][:200]  if len(risks)>2  else '',
    ]

    if TRACKER.exists():
        wb = openpyxl.load_workbook(TRACKER)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Tracker"
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.row_dimensions[1].height = 20

    next_row = ws.max_row + 1
    is_alt = (next_row % 2 == 0)
    for col, val in enumerate(new_row, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if is_alt:
            cell.fill = ALT_FILL

    widths = [18,30,12,20,15,20,14,14,10,10,10,18,14,14,50,50,50,50,50,50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'
    wb.save(TRACKER)
    log.info(f"XLSX  → {TRACKER}  (row {next_row})")
    return TRACKER


# ══════════════════════════════════════════════════════════════════════════════
# 5. PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def process_pdf(pdf_path: str):
    data = parse_report(pdf_path)
    if not data:
        raise ValueError(f"Failed to parse {pdf_path}")
    stem = Path(pdf_path).stem
    return data, write_json(data, stem), write_txt(data, stem), write_tracker(data)

def batch_process(reports_dir: str = str(REPORTS_DIR)) -> list:
    pdfs = list(Path(reports_dir).glob("*.pdf")) + list(Path(reports_dir).glob("*.PDF"))
    if not pdfs:
        log.warning(f"No PDF files found in {reports_dir}")
        return []
    results = []
    for pdf in pdfs:
        try:
            data, *_ = process_pdf(str(pdf))
            results.append(data)
            print(f"  ✓ {pdf.name}")
        except Exception as e:
            log.error(f"  ✗ {pdf.name}: {e}")
    return results


# ══════════════════════════════════════════════════════════════════════════════
# 6. DEMO
# ══════════════════════════════════════════════════════════════════════════════

DEMO_TEXT = """
中信證券
研究部

分析師：張三  CFA
日期：2026-03-15

台積電（2330.TW）深度研究報告

投資評級：買入
目標價：NT$1,200

核心觀點
① AI算力需求帶動CoWoS封裝訂單爆發，2026年產能已全部售磬。
② 2nm製程良率超預期，Q3正式開始量產，蘋果已確認獨家首批。
③ 海外擴產策略分散地緣政治風險，美國廠毛利率回穩至53%以上。

EPS 2026E：NT$45.2
P/E 2026E：26.5x
ROE：25.3%
毛利率：53.5%
營收 2026E：NT$3.2兆

風險提示
• 地緣政治升溫可能衝擊台灣半導體供應鏈
• 客戶集中度過高（蘋果佔比 ~25%），訂單波動影響大
• 美國廠建置成本超支，拖累集團毛利率
"""

def run_demo():
    print("\n🧪 Running DEMO with synthetic report text...\n")
    demo_stem = "demo_TSMC_2026Q1"
    data = {
        "source_file":    f"{demo_stem}.pdf",
        "parsed_at":      datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "date":           extract_date(DEMO_TEXT),
        "institution":    extract_institution(DEMO_TEXT) or "中信證券",
        "analyst":        extract_analyst(DEMO_TEXT),
        "ticker":         extract_ticker(DEMO_TEXT),
        "rating":         extract_rating(DEMO_TEXT),
        "target_price":   extract_target_price(DEMO_TEXT),
        "metrics":        extract_metrics(DEMO_TEXT),
        "thesis":         extract_thesis(DEMO_TEXT),
        "risks":          extract_risks(DEMO_TEXT),
        "raw_text_chars": len(DEMO_TEXT),
    }

    json_path = write_json(data, demo_stem)
    txt_path  = write_txt(data, demo_stem)
    xls_path  = write_tracker(data)

    print("=" * 55)
    print("  PARSED DATA (JSON)")
    print("=" * 55)
    print(json.dumps(data, ensure_ascii=False, indent=2))
    print("\n" + "=" * 55)
    print("  TELEGRAM SUMMARY")
    print("=" * 55)
    print(txt_path.read_text(encoding='utf-8'))
    print("\n" + "=" * 55)
    print("  FILES WRITTEN")
    print("=" * 55)
    print(f"  JSON : {json_path}")
    print(f"  TXT  : {txt_path}")
    print(f"  XLSX : {xls_path}")
    print()


# ══════════════════════════════════════════════════════════════════════════════
# 7. CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    args = sys.argv[1:]
    if not args or args[0] == '--demo':
        run_demo()
    elif args[0] == '--batch':
        directory = args[1] if len(args) > 1 else str(REPORTS_DIR)
        print(f"\nBatch processing PDFs in: {directory}")
        results = batch_process(directory)
        print(f"\nProcessed {len(results)} report(s).")
    else:
        for pdf_arg in args:
            if not Path(pdf_arg).exists():
                print(f"File not found: {pdf_arg}"); continue
            data, json_p, txt_p, xls_p = process_pdf(pdf_arg)
            print(txt_p.read_text(encoding='utf-8'))
            print(f"\nJSON: {json_p}\nXLSX: {xls_p}")
